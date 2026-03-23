import csv
import io
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from django.http import HttpResponse

# Design system brand colors (from masterpage.html :root)
_XLSX_HEADER_BG = "EAF7FB"  # --tp-color-brand-050
_XLSX_HEADER_FG = "0B4A61"  # --tp-color-brand-900
_XLSX_ALT_ROW   = "F4FAF7"  # --tp-color-surface-soft


def _normalize_filename(value, fallback="export"):
    text = str(value or "").strip()
    if not text:
        return fallback
    text = re.sub(r"[^\w.-]+", "-", text, flags=re.ASCII).strip("-._")
    return text or fallback


def csv_response(filename_stem, headers, rows):
    filename = _normalize_filename(filename_stem)
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(["" if value is None else value for value in row])
    return response


def xlsx_sections_response(filename_stem, sections):
    """
    Exports multi-section data as a formatted Excel workbook.
    Each section becomes a separate worksheet with styled headers.
    """
    filename = _normalize_filename(filename_stem)

    header_font = Font(bold=True, color=_XLSX_HEADER_FG)
    header_fill = PatternFill("solid", fgColor=_XLSX_HEADER_BG)
    alt_fill    = PatternFill("solid", fgColor=_XLSX_ALT_ROW)
    center_align = Alignment(vertical="center")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for section in sections:
        title   = str(section.get("title") or "Data").strip() or "Data"
        headers = list(section.get("headers") or [])
        rows    = list(section.get("rows") or [])

        ws = wb.create_sheet(title=title[:31])

        col_widths = [max(len(str(h)), 6) for h in headers]

        if headers:
            ws.append(headers)
            for cell in ws[1]:
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center_align
            ws.row_dimensions[1].height = 22
            ws.freeze_panes = "A2"

        for row_num, row in enumerate(rows, 2):
            normalized = ["" if v is None else v for v in row]
            ws.append(normalized)
            if row_num % 2 == 0:
                for cell in ws[row_num]:
                    cell.fill = alt_fill
            for col_idx, val in enumerate(normalized):
                if col_idx < len(col_widths):
                    col_widths[col_idx] = max(col_widths[col_idx], min(len(str(val)), 64))

        for col_idx, width in enumerate(col_widths, 1):
            col_letter = ws.cell(1, col_idx).column_letter
            ws.column_dimensions[col_letter].width = width + 3

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response = HttpResponse(buffer.read(), content_type=mime)
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response
